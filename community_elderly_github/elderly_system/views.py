"""视图层"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Count, Q, Case, When, BooleanField, Value
from django.utils import timezone
from django.contrib import messages
from django.core.paginator import Paginator
import json
import qrcode
import io
import os
from datetime import datetime, timedelta

from .models import User, Activity, Registration, Announcement, ActivityStats, News


# 工具函数

def json_response(data=None, msg='success', code=200):
    return JsonResponse({'code': code, 'msg': msg, 'data': data or {}})


def admin_required(func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/login/')
        if request.user.role != 'admin':
            return JsonResponse({'code': 403, 'msg': '无权限'})
        return func(request, *args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper


def elder_required(func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('/login/')
        if request.user.role not in ('elder', 'admin'):
            return JsonResponse({'code': 403, 'msg': '无权限'})
        return func(request, *args, **kwargs)
    wrapper.__name__ = func.__name__
    return wrapper


# 认证模块

def login_view(request):
    if request.user.is_authenticated:
        if request.user.role == 'admin':
            return redirect('/admin_dashboard/')
        return redirect('/user_dashboard/')
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        user = authenticate(request, username=username, password=password)
        if user and user.is_active:
            login(request, user)
            if user.role == 'admin':
                return redirect('/admin_dashboard/')
            return redirect('/user_dashboard/')
        else:
            return render(request, 'login.html', {'error': '用户名或密码错误，请重试'})
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    return redirect('/login/')


def register_view(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        real_name = request.POST.get('real_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        age = request.POST.get('age', '').strip()
        area = request.POST.get('area', '').strip()
        if not username:
            return render(request, 'register.html', {'error': '用户名不能为空'})
        if not password:
            return render(request, 'register.html', {'error': '密码不能为空'})
        if len(username) > 150:
            return render(request, 'register.html', {'error': '用户名过长，请控制在150字符以内'})
        if age and (not age.isdigit() or int(age) < 0 or int(age) > 150):
            return render(request, 'register.html', {'error': '年龄格式不正确，请填写0-150之间的数字'})
        if User.objects.filter(username=username).exists():
            return render(request, 'register.html', {'error': '用户名已存在'})
        if phone and User.objects.filter(phone=phone).exists():
            return render(request, 'register.html', {'error': '手机号已注册'})
        user = User.objects.create_user(
            username=username,
            password=password,
            first_name=real_name,
            phone=phone,
            age=int(age) if age else None,
            area=area,
            role='elder'
        )
        login(request, user)
        return redirect('/user_dashboard/')
    return render(request, 'register.html')


# 首页跳转

@login_required
def index(request):
    if request.user.role == 'admin':
        return redirect('/admin_dashboard/')
    return redirect('/user_dashboard/')


# 管理员端

@admin_required
def admin_dashboard(request):
    now = timezone.now()
    total_users = User.objects.filter(role='elder').count()
    total_activities = Activity.objects.count()
    open_activities = Activity.objects.filter(
        status__in=['open', 'ongoing']
    ).exclude(
        status='open', register_deadline__lt=now
    ).count()
    total_registrations = Registration.objects.exclude(status='cancelled').count()
    trend_data = []
    today = timezone.now().date()
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        count = Registration.objects.filter(
            register_time__date=day
        ).exclude(status='cancelled').count()
        trend_data.append({'date': day.strftime('%m/%d'), 'count': count})
    announcements = Announcement.objects.order_by('-publish_time')[:5]
    recent_activities = Activity.objects.order_by('-created_at')[:5]
    context = {
        'total_users': total_users,
        'total_activities': total_activities,
        'open_activities': open_activities,
        'total_registrations': total_registrations,
        'trend_data': json.dumps(trend_data, ensure_ascii=False),
        'announcements': announcements,
        'recent_activities': recent_activities,
    }
    return render(request, 'admin/dashboard.html', context)


@admin_required
def admin_users(request):
    search = request.GET.get('search', '')
    area = request.GET.get('area', '')
    qs = User.objects.filter(role='elder')
    if search:
        qs = qs.filter(Q(username__icontains=search) | Q(first_name__icontains=search) | Q(phone__icontains=search))
    if area:
        qs = qs.filter(area__icontains=area)
    qs = qs.order_by('-created_at')
    paginator = Paginator(qs, 10)
    page = paginator.get_page(request.GET.get('page', 1))
    areas = User.objects.filter(role='elder').exclude(area='').exclude(area=None).values_list('area', flat=True).distinct()
    return render(request, 'admin/users.html', {'page': page, 'areas': areas, 'search': search, 'selected_area': area})


@admin_required
def admin_user_toggle(request, user_id):
    user = get_object_or_404(User, id=user_id, role='elder')
    user.is_active = not user.is_active
    user.save()
    return json_response({'is_active': user.is_active}, msg='操作成功')


@admin_required
def admin_activities(request):
    status = request.GET.get('status', '')
    act_type = request.GET.get('type', '')
    search = request.GET.get('search', '')
    qs = Activity.objects.all()
    if status:
        qs = qs.filter(status=status)
    if act_type:
        qs = qs.filter(activity_type=act_type)
    if search:
        qs = qs.filter(Q(title__icontains=search) | Q(location__icontains=search))
    paginator = Paginator(qs, 8)
    page = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'admin/activities.html', {
        'page': page,
        'status_choices': Activity.STATUS_CHOICES,
        'type_choices': Activity.TYPE_CHOICES,
        'current_status': status,
        'current_type': act_type,
        'search': search,
    })


@admin_required
def admin_activity_create(request):
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        location = request.POST.get('location', '').strip()
        activity_type = request.POST.get('activity_type', 'other')
        start_time = request.POST.get('start_time')
        end_time = request.POST.get('end_time')
        register_deadline = request.POST.get('register_deadline')
        max_participants = request.POST.get('max_participants', 50)
        target_group = request.POST.get('target_group', '')
        status = request.POST.get('status', 'draft')
        activity = Activity.objects.create(
            title=title,
            description=description,
            location=location,
            activity_type=activity_type,
            start_time=start_time,
            end_time=end_time,
            register_deadline=register_deadline,
            max_participants=int(max_participants),
            target_group=target_group,
            status=status,
            creator=request.user,
        )
        ActivityStats.objects.create(activity=activity)
        if request.FILES.get('poster'):
            activity.poster = request.FILES['poster']
            activity.save()
        messages.success(request, f'活动「{title}」创建成功！')
        return redirect('/admin/activities/')
    return render(request, 'admin/activity_form.html', {
        'status_choices': Activity.STATUS_CHOICES,
        'type_choices': Activity.TYPE_CHOICES,
        'action': '创建',
    })


@admin_required
def admin_activity_edit(request, act_id):
    activity = get_object_or_404(Activity, id=act_id)
    if request.method == 'POST':
        from datetime import datetime as dt
        activity.title = request.POST.get('title', activity.title).strip()
        activity.description = request.POST.get('description', activity.description)
        activity.location = request.POST.get('location', activity.location).strip()
        activity.activity_type = request.POST.get('activity_type', activity.activity_type)
        start_time_str = request.POST.get('start_time', '')
        end_time_str = request.POST.get('end_time', '')
        deadline_str = request.POST.get('register_deadline', '')
        if start_time_str:
            try:
                activity.start_time = dt.fromisoformat(start_time_str)
            except (ValueError, TypeError):
                messages.error(request, '开始时间格式不正确')
                return render(request, 'admin/activity_form.html', {'activity': activity, 'status_choices': Activity.STATUS_CHOICES, 'type_choices': Activity.TYPE_CHOICES, 'action': '编辑'})
        if end_time_str:
            try:
                activity.end_time = dt.fromisoformat(end_time_str)
            except (ValueError, TypeError):
                messages.error(request, '结束时间格式不正确')
                return render(request, 'admin/activity_form.html', {'activity': activity, 'status_choices': Activity.STATUS_CHOICES, 'type_choices': Activity.TYPE_CHOICES, 'action': '编辑'})
        if deadline_str:
            try:
                activity.register_deadline = dt.fromisoformat(deadline_str)
            except (ValueError, TypeError):
                messages.error(request, '报名截止时间格式不正确')
                return render(request, 'admin/activity_form.html', {'activity': activity, 'status_choices': Activity.STATUS_CHOICES, 'type_choices': Activity.TYPE_CHOICES, 'action': '编辑'})
        if activity.start_time and activity.end_time and activity.end_time <= activity.start_time:
            messages.error(request, '结束时间必须晚于开始时间')
            return render(request, 'admin/activity_form.html', {'activity': activity, 'status_choices': Activity.STATUS_CHOICES, 'type_choices': Activity.TYPE_CHOICES, 'action': '编辑'})
        if activity.start_time and activity.register_deadline and activity.register_deadline >= activity.start_time:
            messages.error(request, '报名截止时间必须早于活动开始时间')
            return render(request, 'admin/activity_form.html', {'activity': activity, 'status_choices': Activity.STATUS_CHOICES, 'type_choices': Activity.TYPE_CHOICES, 'action': '编辑'})
        activity.max_participants = int(request.POST.get('max_participants', activity.max_participants))
        activity.target_group = request.POST.get('target_group', activity.target_group)
        activity.status = request.POST.get('status', activity.status)
        if request.FILES.get('poster'):
            activity.poster = request.FILES['poster']
        activity.save()
        messages.success(request, '活动信息更新成功！')
        return redirect('/admin/activities/')
    return render(request, 'admin/activity_form.html', {
        'activity': activity,
        'status_choices': Activity.STATUS_CHOICES,
        'type_choices': Activity.TYPE_CHOICES,
        'action': '编辑',
    })


@admin_required
@require_POST
def admin_activity_delete(request, act_id):
    activity = get_object_or_404(Activity, id=act_id)
    title = activity.title
    if activity.poster and os.path.exists(activity.poster.path):
        os.remove(activity.poster.path)
    for reg in Registration.objects.filter(activity=activity):
        if reg.qr_code and os.path.exists(reg.qr_code.path):
            try:
                os.remove(reg.qr_code.path)
            except OSError:
                pass
    activity.delete()
    messages.success(request, f'活动「{title}」已删除')
    return redirect('/admin/activities/')


@admin_required
def admin_registrations(request):
    act_id = request.GET.get('activity_id', '')
    qs = Registration.objects.select_related('user', 'activity')
    if act_id:
        qs = qs.filter(activity_id=act_id)
    paginator = Paginator(qs, 15)
    page = paginator.get_page(request.GET.get('page', 1))
    activities = Activity.objects.filter(status__in=['open', 'ongoing', 'finished']).values('id', 'title')
    return render(request, 'admin/registrations.html', {
        'page': page,
        'activities': activities,
        'current_activity': act_id,
    })


@admin_required
def admin_sign_in(request, reg_id):
    reg = get_object_or_404(Registration, id=reg_id)
    if reg.status == 'registered':
        reg.status = 'signed_in'
        reg.sign_in_time = timezone.now()
        reg.save()
        stats, _ = ActivityStats.objects.get_or_create(activity=reg.activity)
        stats.recalculate()
    return json_response(msg='签到成功')


@admin_required
def admin_announcements(request):
    qs = Announcement.objects.select_related('publisher').order_by('-is_top', '-publish_time')
    paginator = Paginator(qs, 10)
    page = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'admin/announcements.html', {'page': page})


@admin_required
def admin_announcement_create(request):
    if request.method == 'POST':
        ann = Announcement.objects.create(
            title=request.POST.get('title', '').strip(),
            content=request.POST.get('content', '').strip(),
            ann_type=request.POST.get('ann_type', 'notice'),
            is_top=request.POST.get('is_top') == 'on',
            publisher=request.user,
        )
        messages.success(request, '公告发布成功！')
        return redirect('/admin/announcements/')
    return render(request, 'admin/announcement_form.html', {
        'type_choices': Announcement.TYPE_CHOICES,
        'action': '发布',
    })


@admin_required
@require_POST
def admin_announcement_delete(request, ann_id):
    ann = get_object_or_404(Announcement, id=ann_id)
    if hasattr(ann, 'cover') and ann.cover and os.path.exists(ann.cover.path):
        try:
            os.remove(ann.cover.path)
        except OSError:
            pass
    ann.delete()
    messages.success(request, '公告已删除')
    return redirect('/admin/announcements/')


@admin_required
def admin_statistics(request):
    type_stats = Activity.objects.values('activity_type').annotate(cnt=Count('id')).order_by('-cnt')
    type_labels = [dict(Activity.TYPE_CHOICES).get(t['activity_type'], t['activity_type']) for t in type_stats]
    type_counts = [t['cnt'] for t in type_stats]
    today = timezone.now().date()
    weekly = []
    for i in range(11, -1, -1):
        start = today - timedelta(weeks=i + 1)
        end = today - timedelta(weeks=i)
        cnt = Registration.objects.filter(register_time__date__gte=start, register_time__date__lt=end).exclude(status='cancelled').count()
        weekly.append({'week': f'第{12-i}周', 'count': cnt})
    top_activities = ActivityStats.objects.select_related('activity').order_by('-sign_in_rate')[:5]
    area_stats = User.objects.filter(role='elder').exclude(area=None).exclude(area='').values('area').annotate(cnt=Count('id')).order_by('-cnt')[:8]
    context = {
        'type_labels': json.dumps(type_labels, ensure_ascii=False),
        'type_counts': json.dumps(type_counts),
        'weekly_data': json.dumps(weekly, ensure_ascii=False),
        'top_activities': top_activities,
        'area_stats': area_stats,
        'total_users': User.objects.filter(role='elder').count(),
        'total_activities': Activity.objects.count(),
        'finished_activities': Activity.objects.filter(status='finished').count(),
        'total_sign_ins': Registration.objects.filter(status='signed_in').count(),
    }
    return render(request, 'admin/statistics.html', context)


@admin_required
def export_registrations(request, act_id):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    activity = get_object_or_404(Activity, id=act_id)
    registrations = Registration.objects.filter(activity=activity).select_related('user').exclude(status='cancelled')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '报名名单'
    headers = ['序号', '姓名', '用户名', '手机号', '年龄', '居住片区', '报名时间', '签到状态', '签到时间']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    for i, reg in enumerate(registrations, 1):
        ws.append([
            i,
            reg.user.first_name or reg.user.username,
            reg.user.username,
            reg.user.phone or '',
            reg.user.age or '',
            reg.user.area or '',
            reg.register_time.strftime('%Y-%m-%d %H:%M') if reg.register_time else '',
            '已签到' if reg.status == 'signed_in' else '未签到',
            reg.sign_in_time.strftime('%Y-%m-%d %H:%M') if reg.sign_in_time else '',
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{activity.title}_报名名单.xlsx"'
    wb.save(response)
    return response


@admin_required
def admin_statistics_export(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.db.models.functions import TruncMonth, TruncQuarter
    wb = openpyxl.Workbook()
    title_font = Font(name='微软雅黑', size=16, bold=True, color='1F2937')
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    sub_header_font = Font(name='微软雅黑', size=11, bold=True, color='1F2937')
    cell_font = Font(name='微软雅黑', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    def style_header_row(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

    def style_cell(ws, row, col, align=center):
        cell = ws.cell(row=row, column=col)
        cell.font = cell_font
        cell.alignment = align
        cell.border = thin_border
        return cell

    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    act_type = request.GET.get('act_type', '')
    activities = Activity.objects.all()
    registrations = Registration.objects.exclude(status='cancelled')
    users = User.objects.filter(role='elder')
    if start_date:
        try:
            sd = datetime.strptime(start_date, '%Y-%m-%d')
            registrations = registrations.filter(register_time__gte=sd)
            activities = activities.filter(start_time__gte=sd)
        except ValueError:
            pass
    if end_date:
        try:
            ed = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)
            registrations = registrations.filter(register_time__lt=ed)
            activities = activities.filter(start_time__lt=ed)
        except ValueError:
            pass
    if act_type:
        activities = activities.filter(activity_type=act_type)
        registrations = registrations.filter(activity__activity_type=act_type)

    # Sheet 1: 活动参与数据统计
    ws1 = wb.active
    ws1.title = '活动参与数据统计'
    ws1.merge_cells('A1:H1')
    title_cell = ws1.cell(row=1, column=1, value='社区老年活动管理系统 — 活动参与数据统计报表')
    title_cell.font = title_font
    title_cell.alignment = center
    ws1.merge_cells('A2:H2')
    filter_desc = f"统计区间: {start_date or '全部'} 至 {end_date or '全部'}  |  "
    filter_desc += f"活动类型: {dict(Activity.TYPE_CHOICES).get(act_type, '全部')}"
    ws1.cell(row=2, column=1, value=filter_desc).font = Font(name='微软雅黑', size=10, color='6B7280')
    row = 4
    ws1.merge_cells(f'A{row}:H{row}')
    ws1.cell(row=row, column=1, value='一、概要统计').font = Font(name='微软雅黑', size=13, bold=True, color='1F2937')
    row += 1
    headers = ['总活动数', '总报名人次', '已签到人次', '总签到率', '异常未签到人次', '异常率']
    for c, h in enumerate(headers, 1):
        ws1.cell(row=row, column=c, value=h)
    style_header_row(ws1, row, len(headers))
    row += 1
    total_acts = activities.count()
    total_regs = registrations.count()
    total_signs = registrations.filter(status='signed_in').count()
    total_abnormal = registrations.filter(status='registered').count()
    sign_rate = f'{total_signs / total_regs * 100:.1f}%' if total_regs > 0 else '0%'
    abnormal_rate = f'{total_abnormal / total_regs * 100:.1f}%' if total_regs > 0 else '0%'
    for c, v in enumerate([total_acts, total_regs, total_signs, sign_rate, total_abnormal, abnormal_rate], 1):
        style_cell(ws1, row, c).value = v
    row += 2
    ws1.merge_cells(f'A{row}:H{row}')
    ws1.cell(row=row, column=1, value='二、活动报名签到明细').font = Font(name='微软雅黑', size=13, bold=True, color='1F2937')
    row += 1
    headers2 = ['序号', '活动名称', '活动类型', '报名人数', '签到人数', '签到率', '异常人数', '异常率']
    for c, h in enumerate(headers2, 1):
        ws1.cell(row=row, column=c, value=h)
    style_header_row(ws1, row, len(headers2))
    row += 1
    for i, act in enumerate(activities, 1):
        regs = Registration.objects.filter(activity=act).exclude(status='cancelled')
        reg_count = regs.count()
        sign_count = regs.filter(status='signed_in').count()
        abn_count = regs.filter(status='registered').count()
        s_rate = f'{sign_count / reg_count * 100:.1f}%' if reg_count > 0 else '0%'
        a_rate = f'{abn_count / reg_count * 100:.1f}%' if reg_count > 0 else '0%'
        values = [i, act.title, act.get_activity_type_display(), reg_count, sign_count, s_rate, abn_count, a_rate]
        for c, v in enumerate(values, 1):
            style_cell(ws1, row, c).value = v
        row += 1
    row += 1
    ws1.merge_cells(f'A{row}:H{row}')
    ws1.cell(row=row, column=1, value='三、参与用户年龄分布').font = Font(name='微软雅黑', size=13, bold=True, color='1F2937')
    row += 1
    age_headers = ['年龄段', '参与人次', '占比']
    for c, h in enumerate(age_headers, 1):
        ws1.cell(row=row, column=c, value=h)
    style_header_row(ws1, row, len(age_headers))
    row += 1
    participant_ids = registrations.values_list('user_id', flat=True).distinct()
    participants = users.filter(id__in=participant_ids)
    age_ranges = [('60岁以下', 0, 60), ('60-65岁', 60, 65), ('66-70岁', 66, 70),
                  ('71-75岁', 71, 75), ('76-80岁', 76, 80), ('80岁以上', 81, 200)]
    for label, lo, hi in age_ranges:
        cnt = participants.filter(age__gte=lo, age__lte=hi).count()
        rate = f'{cnt / participants.count() * 100:.1f}%' if participants.count() > 0 else '0%'
        for c, v in enumerate([label, cnt, rate], 1):
            style_cell(ws1, row, c).value = v
        row += 1
    row += 1
    ws1.merge_cells(f'A{row}:H{row}')
    ws1.cell(row=row, column=1, value='四、参与用户片区分布').font = Font(name='微软雅黑', size=13, bold=True, color='1F2937')
    row += 1
    for c, h in enumerate(['居住片区', '参与人次', '占比'], 1):
        ws1.cell(row=row, column=c, value=h)
    style_header_row(ws1, row, 3)
    row += 1
    area_counts = participants.values('area').annotate(cnt=Count('id')).order_by('-cnt')
    total_area = participants.count()
    for item in area_counts:
        label = item['area'] or '未填写'
        cnt = item['cnt']
        rate = f'{cnt / total_area * 100:.1f}%' if total_area > 0 else '0%'
        for c, v in enumerate([label, cnt, rate], 1):
            style_cell(ws1, row, c).value = v
        row += 1
    for c in range(1, 9):
        ws1.column_dimensions[get_column_letter(c)].width = 18

    # Sheet 2: 用户活跃数据统计
    ws2 = wb.create_sheet('用户活跃数据统计')
    ws2.merge_cells('A1:F1')
    ws2.cell(row=1, column=1, value='社区老年活动管理系统 — 用户活跃数据统计报表').font = title_font
    ws2.cell(row=1, column=1).alignment = center
    ws2.merge_cells('A2:F2')
    ws2.cell(row=2, column=1, value=filter_desc).font = Font(name='微软雅黑', size=10, color='6B7280')
    row2 = 4
    ws2.merge_cells(f'A{row2}:F{row2}')
    ws2.cell(row=row2, column=1, value='一、月度参与频次统计').font = Font(name='微软雅黑', size=13, bold=True, color='1F2937')
    row2 += 1
    for c, h in enumerate(['月份', '参与人次', '活跃用户数', '人均参与次数'], 1):
        ws2.cell(row=row2, column=c, value=h)
    style_header_row(ws2, row2, 4)
    row2 += 1
    monthly = registrations.annotate(month=TruncMonth('register_time')).values('month').annotate(
        cnt=Count('id'), users=Count('user_id', distinct=True)
    ).order_by('-month')[:12]
    for m in reversed(monthly):
        month_str = m['month'].strftime('%Y年%m月') if m['month'] else '未知'
        avg = f'{m["cnt"] / m["users"]:.1f}' if m['users'] > 0 else '0'
        for c, v in enumerate([month_str, m['cnt'], m['users'], avg], 1):
            style_cell(ws2, row2, c).value = v
        row2 += 1
    row2 += 1
    ws2.merge_cells(f'A{row2}:F{row2}')
    ws2.cell(row=row2, column=1, value='二、片区用户参与度差异').font = Font(name='微软雅黑', size=13, bold=True, color='1F2937')
    row2 += 1
    for c, h in enumerate(['居住片区', '用户总数', '活跃用户数', '总参与人次', '人均参与次数', '活跃率'], 1):
        ws2.cell(row=row2, column=c, value=h)
    style_header_row(ws2, row2, 6)
    row2 += 1
    area_data = users.exclude(area='').exclude(area=None).values('area').annotate(
        total=Count('id')
    ).order_by('-total')
    for ad in area_data:
        area_users = users.filter(area=ad['area'])
        area_user_ids = area_users.values_list('id', flat=True)
        area_regs = registrations.filter(user_id__in=area_user_ids)
        active_users = area_regs.values('user_id').distinct().count()
        total_parts = area_regs.count()
        avg_parts = f'{total_parts / active_users:.1f}' if active_users > 0 else '0'
        active_rate = f'{active_users / ad["total"] * 100:.1f}%' if ad['total'] > 0 else '0%'
        for c, v in enumerate([ad['area'], ad['total'], active_users, total_parts, avg_parts, active_rate], 1):
            style_cell(ws2, row2, c).value = v
        row2 += 1
    row2 += 1
    ws2.merge_cells(f'A{row2}:F{row2}')
    ws2.cell(row=row2, column=1, value='三、热门活动类型覆盖范围').font = Font(name='微软雅黑', size=13, bold=True, color='1F2937')
    row2 += 1
    for c, h in enumerate(['活动类型', '活动数量', '参与人次', '覆盖片区数', '覆盖用户数'], 1):
        ws2.cell(row=row2, column=c, value=h)
    style_header_row(ws2, row2, 5)
    row2 += 1
    type_data = activities.values('activity_type').annotate(cnt=Count('id'), parts=Count('registrations')).order_by('-parts')
    for td in type_data:
        type_acts = activities.filter(activity_type=td['activity_type'])
        act_ids = type_acts.values_list('id', flat=True)
        type_regs = registrations.filter(activity_id__in=act_ids)
        type_user_ids = type_regs.values_list('user_id', flat=True).distinct()
        area_count = users.filter(id__in=type_user_ids).exclude(area='').exclude(area=None).values('area').distinct().count()
        user_count = type_user_ids.count()
        for c, v in enumerate([dict(Activity.TYPE_CHOICES).get(td['activity_type'], td['activity_type']),
                                td['cnt'], td['parts'], area_count, user_count], 1):
            style_cell(ws2, row2, c).value = v
        row2 += 1
    for c in range(1, 7):
        ws2.column_dimensions[get_column_letter(c)].width = 22
    filename = f'统计报表_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# 用户端

@elder_required
def user_dashboard(request):
    user = request.user
    now = timezone.now()
    announcements = Announcement.objects.order_by('-is_top', '-publish_time')[:5]
    open_activities = Activity.objects.filter(
        status='open',
        register_deadline__gt=now
    ).order_by('start_time')[:6]
    my_registrations = Registration.objects.filter(
        user=user
    ).exclude(status='cancelled').select_related('activity').filter(
        activity__end_time__gte=now
    ).order_by('activity__start_time')[:3]
    total_my_regs = Registration.objects.filter(
        user=user
    ).exclude(status='cancelled').count()
    total_participated = Registration.objects.filter(
        user=user, status='signed_in'
    ).count()
    context = {
        'user': user,
        'announcements': announcements,
        'open_activities': open_activities,
        'my_registrations': my_registrations,
        'total_my_regs': total_my_regs,
        'total_participated': total_participated,
    }
    return render(request, 'user/dashboard.html', context)


@elder_required
def user_activities(request):
    act_type = request.GET.get('type', '')
    search = request.GET.get('search', '')
    now = timezone.now()
    qs = Activity.objects.filter(status__in=['open', 'ongoing', 'finished'])
    qs = qs.annotate(
        is_reg_closed=Case(
            When(status='open', register_deadline__lt=now, then=Value(True)),
            default=Value(False),
            output_field=BooleanField(),
        )
    )
    if act_type:
        qs = qs.filter(activity_type=act_type)
    if search:
        qs = qs.filter(Q(title__icontains=search) | Q(location__icontains=search))
    qs = qs.order_by('-start_time')
    paginator = Paginator(qs, 9)
    page = paginator.get_page(request.GET.get('page', 1))
    my_reg_ids = set(Registration.objects.filter(
        user=request.user
    ).exclude(status='cancelled').values_list('activity_id', flat=True))
    return render(request, 'user/activities.html', {
        'page': page,
        'type_choices': Activity.TYPE_CHOICES,
        'current_type': act_type,
        'search': search,
        'my_reg_ids': my_reg_ids,
        'now': now,
    })


@elder_required
def user_activity_detail(request, act_id):
    activity = get_object_or_404(Activity, id=act_id)
    now = timezone.now()
    is_registration_closed = activity.register_deadline and activity.register_deadline < now
    is_activity_finished = activity.end_time and activity.end_time < now
    is_activity_ongoing = (activity.start_time and activity.start_time <= now
                           and activity.end_time and activity.end_time >= now)
    is_registered = Registration.objects.filter(
        user=request.user, activity=activity
    ).exclude(status='cancelled').exists()
    my_reg = None
    if is_registered:
        my_reg = Registration.objects.filter(user=request.user, activity=activity).first()
    return render(request, 'user/activity_detail.html', {
        'activity': activity,
        'is_registered': is_registered,
        'my_reg': my_reg,
        'is_registration_closed': is_registration_closed,
        'is_activity_finished': is_activity_finished,
        'is_activity_ongoing': is_activity_ongoing,
    })


@elder_required
@require_POST
def user_register_activity(request, act_id):
    activity = get_object_or_404(Activity, id=act_id)
    user = request.user
    if activity.status != 'open':
        return json_response(msg='该活动暂不接受报名', code=400)
    if activity.register_deadline and activity.register_deadline < timezone.now():
        return json_response(msg='报名已截止', code=400)
    if activity.is_full:
        return json_response(msg='报名人数已满', code=400)
    if Registration.objects.filter(user=user, activity=activity).exclude(status='cancelled').exists():
        return json_response(msg='您已报名该活动', code=400)
    reg = Registration.objects.create(user=user, activity=activity, status='registered')
    try:
        qr_data = f"sign_in:{reg.id}:{user.id}:{activity.id}"
        qr_img = qrcode.make(qr_data)
        img_io = io.BytesIO()
        qr_img.save(img_io, 'PNG')
        img_io.seek(0)
        from django.core.files.base import ContentFile
        reg.qr_code.save(f'qr_{reg.id}.png', ContentFile(img_io.read()), save=True)
    except Exception:
        pass
    from django.db.models import F
    activity.current_participants = F('current_participants') + 1
    activity.save(update_fields=['current_participants'])
    stats, _ = ActivityStats.objects.get_or_create(activity=activity)
    stats.recalculate()
    return json_response(msg='报名成功！')


@elder_required
@require_POST
def user_cancel_registration(request, act_id):
    activity = get_object_or_404(Activity, id=act_id)
    reg = get_object_or_404(Registration, user=request.user, activity=activity)
    if reg.status == 'signed_in':
        return json_response(msg='已签到，无法取消报名', code=400)
    if activity.register_deadline and activity.register_deadline < timezone.now():
        return json_response(msg='报名已截止，无法取消报名', code=400)
    if activity.end_time and activity.end_time < timezone.now():
        return json_response(msg='活动已结束，无法取消报名', code=400)
    reg.status = 'cancelled'
    reg.save()
    from django.db.models import F
    activity.current_participants = F('current_participants') - 1
    activity.save(update_fields=['current_participants'])
    return json_response(msg='已取消报名')


@elder_required
def user_my_registrations(request):
    status = request.GET.get('status', '')
    qs = Registration.objects.filter(user=request.user).select_related('activity')
    if status:
        qs = qs.filter(status=status)
    qs = qs.order_by('-register_time')
    paginator = Paginator(qs, 10)
    page = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'user/my_registrations.html', {
        'page': page,
        'status_choices': Registration.STATUS_CHOICES,
        'current_status': status,
        'now': timezone.now(),
    })


@elder_required
def user_profile(request):
    if request.method == 'POST':
        user = request.user
        user.first_name = request.POST.get('real_name', user.first_name)
        user.phone = request.POST.get('phone', user.phone)
        user.age = request.POST.get('age') or user.age
        user.area = request.POST.get('area', user.area)
        user.family_phone = request.POST.get('family_phone', user.family_phone)
        if request.FILES.get('avatar'):
            user.avatar = request.FILES['avatar']
        user.save()
        messages.success(request, '个人信息更新成功！')
        return redirect('/user/profile/')
    return render(request, 'user/profile.html', {'user': request.user})


@elder_required
def user_announcements(request):
    qs = Announcement.objects.order_by('-is_top', '-publish_time')
    paginator = Paginator(qs, 10)
    page = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'user/announcements.html', {'page': page})


# 新闻资讯模块

@admin_required
def admin_news(request):
    qs = News.objects.all()
    category = request.GET.get('category', '')
    if category:
        qs = qs.filter(category=category)
    paginator = Paginator(qs, 10)
    page = paginator.get_page(request.GET.get('page', 1))
    categories = News.CATEGORY_CHOICES
    return render(request, 'admin/news.html', {'page': page, 'category': category, 'categories': categories})


@admin_required
def admin_news_create(request):
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        content = request.POST.get('content', '').strip()
        category = request.POST.get('category', 'other')
        source = request.POST.get('source', '').strip()
        is_top = request.POST.get('is_top') == 'on'
        cover = request.FILES.get('cover')
        if not title or not content:
            messages.error(request, '标题和内容不能为空')
            temp_news = News(
                title=title,
                content=content,
                category=category,
                source=source,
                is_top=is_top,
            )
            return render(request, 'admin/news_form.html', {'news': temp_news})
        news = News.objects.create(
            title=title,
            content=content,
            category=category,
            source=source,
            is_top=is_top,
            cover=cover,
            author=request.user,
        )
        messages.success(request, f'新闻《{news.title}》发布成功')
        return redirect('/admin/news/')
    empty_news = News()
    empty_news.source = '社区发布'
    return render(request, 'admin/news_form.html', {'news': empty_news})


@admin_required
def admin_news_edit(request, news_id):
    news = get_object_or_404(News, id=news_id)
    if request.method == 'POST':
        news.title = request.POST.get('title', '').strip()
        news.content = request.POST.get('content', '').strip()
        news.category = request.POST.get('category', 'other')
        news.source = request.POST.get('source', '').strip()
        news.is_top = request.POST.get('is_top') == 'on'
        views_str = request.POST.get('views', '0')
        try:
            news.views = int(views_str)
        except (ValueError, TypeError):
            news.views = 0
        if request.FILES.get('cover'):
            news.cover = request.FILES.get('cover')
        if news.title and news.content:
            news.save()
            messages.success(request, '新闻更新成功')
            return redirect('/admin/news/')
        else:
            messages.error(request, '标题和内容不能为空')
            return render(request, 'admin/news_form.html', {'news': news, 'form_data': request.POST})
    return render(request, 'admin/news_form.html', {'news': news, 'form_data': {}})


@admin_required
@require_POST
def admin_news_delete(request, news_id):
    news = get_object_or_404(News, id=news_id)
    if news.cover and os.path.exists(news.cover.path):
        try:
            os.remove(news.cover.path)
        except OSError:
            pass
    news.delete()
    messages.success(request, '新闻已删除')
    return redirect('/admin/news/')


@elder_required
def user_news(request):
    category = request.GET.get('category', '')
    qs = News.objects.all()
    if category:
        qs = qs.filter(category=category)
    paginator = Paginator(qs, 9)
    page = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'user/news.html', {'page': page, 'category': category})


@elder_required
def user_news_detail(request, news_id):
    news = get_object_or_404(News, id=news_id)
    from django.db.models import F
    News.objects.filter(id=news_id).update(views=F('views') + 1)
    news.refresh_from_db()
    related = News.objects.filter(category=news.category).exclude(id=news_id)[:4]
    return render(request, 'user/news_detail.html', {'news': news, 'related': related})
